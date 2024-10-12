import os
import pandas as pd
import json
from openpyxl import load_workbook

def write_invoices_to_excel(json_file, output_file):
    with open(json_file, 'r', encoding='utf-8') as file:
        invoices = json.load(file)
    
    df = pd.DataFrame(invoices, columns=['invoice_number', 'seller', 'buyer', 'amount', 'transaction_date', 'approval_status','manual_review_reason','img_url'])
    df.columns = ['发票号码', '卖方', '买方', '金额', '交易时间', '审批状态', '转人工审批原因', '图片链接']

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name='发票数据详情')

def write_seller_data_to_excel(json_file, output_file):
    with open(json_file, 'r', encoding='utf-8') as file:
        seller_data = json.load(file)
    
    df = pd.DataFrame(seller_data, columns=['seller', 'total_amount', 'transaction_count', 'customer_category'])
    df.columns = ['卖方', '交易总额', '交易次数', '客户分类']

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name='卖方数据详情')

def write_buyer_data_to_excel(json_file, output_file):
    with open(json_file, 'r', encoding='utf-8') as file:
        buyer_data = json.load(file)
    
    df = pd.DataFrame(buyer_data, columns=['buyer', 'total_amount', 'transaction_frequency', 'customer_category'])
    df.columns = ['买方', '交易总额', '交易频度', '客户分类']

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name='买方数据详情')

def write_transaction_summary_to_excel(json_file, output_file):
    # 读取 json 文件
    with open(json_file, 'r', encoding='utf-8') as file:
        transaction_summary = json.load(file)

    # 转换为 DataFrame
    df = pd.DataFrame(transaction_summary['transaction_summary'].items(), columns=['key', 'value'])
    df['value'] = df['value'].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
    df.columns = ['交易关系汇总', '']

    # 每一行的第一列 转换为对应的中文
    df['交易关系汇总'] = df['交易关系汇总'].map({
        'major_clients': '大客户',
        'clients': '客户',
        'regular_clients': '一般客户',
        'important_suppliers': '重要供应商',
        'suppliers': '供应商',
        'regular_suppliers': '一般供应商',
        'top_buyers_by_purchase_volume': '当前购买量最大的3个买方',
        'top_sellers_by_sales_volume': '当前卖出量最大的3个卖方',
        'most_frequent_transaction_relationship': '最频繁交易关系'
    })

    # 对于最后面的 '最频繁交易关系'，读出它的value，{'client': 'Client1', 'supplier': 'Supplier1'}
    # 把这两个合并成一个字符串，然后写入 DataFrame
    most_frequent_transaction_relationship = transaction_summary['transaction_summary']['most_frequent_transaction_relationship']
    most_frequent_transaction_relationship_str = f"客户：{most_frequent_transaction_relationship['client']} - 供应商：{most_frequent_transaction_relationship['supplier']}"
    df.loc[df['交易关系汇总'] == '最频繁交易关系', '',] = most_frequent_transaction_relationship_str

    # 写入 Excel 文件
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name='交易关系汇总')

def write_invoice_approval_summary_to_excel(json_file, output_file):
    with open(json_file, 'r', encoding='utf-8') as file:
        invoice_approval_summary = json.load(file)

    df = pd.DataFrame(invoice_approval_summary['Invoice Approval Summary'].items(), columns=['key', 'value'])
    df['value'] = df['value'].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
    df.columns = ['发票审批状态汇总', '']

    df['发票审批状态汇总'] = df['发票审批状态汇总'].map({
        'Total Invoices': '总发票数量',
        'Approved Invoices': '通过发票数量',
        'Rejected Invoices': '不通过发票数量',
        'Invoices Sent for Manual Review': '转人工审批发票数量',
        'Approval Status Ratio': '审批状态比例',
        'Maximum Invoice Amount': '发票最大金额',
        'Minimum Invoice Amount': '发票最小金额',
        'Average Invoice Amount': '发票平均金额',
        'Most Common Reason for Manual Review': '最多转人工审批原因',
        'Duplicate Invoice Count': '重复发票张数'
    })

    # 把审批状态比例的值，从字典转换为字符串
    approval_status_ratio = invoice_approval_summary['Invoice Approval Summary']['Approval Status Ratio']
    approval_status_ratio_str = f"{approval_status_ratio['Approved']} / {approval_status_ratio['Rejected']} / {approval_status_ratio['Manual Review']}"
    df.loc[df['发票审批状态汇总'] == '审批状态比例', '',] = approval_status_ratio_str

    # 写入 Excel 文件
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name='发票审批状态汇总')
  

# 指定输入和输出文件名
input_invoice_json_file = './.json/invoice.json'
input_seller_json_file = './.json/seller.json'
input_buyer_json_file = './.json/buyer.json'
input_transaction_summary_json_file = './.json/transaction_summary.json'
input_invoice_approval_summary_json_file = './.json/invoice_approval_summary.json'
output_excel_file = './.xlsx/发票自动化处理统计报表.xlsx'

# 如果文件已存在，则删除
if os.path.exists(output_excel_file):
    os.remove(output_excel_file)

# 创建新的 Excel 文件并添加一个空白 sheet
with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
    pd.DataFrame().to_excel(writer, index=False, sheet_name='z')

# 在新的文件中写入数据
write_invoices_to_excel(input_invoice_json_file, output_excel_file)
write_seller_data_to_excel(input_seller_json_file, output_excel_file)
write_buyer_data_to_excel(input_buyer_json_file, output_excel_file)
write_transaction_summary_to_excel(input_transaction_summary_json_file, output_excel_file)
write_invoice_approval_summary_to_excel(input_invoice_approval_summary_json_file, output_excel_file)

# 删除空白的 'z' sheet
wb = load_workbook(output_excel_file)
if 'z' in wb.sheetnames:
    del wb['z']
wb.save(output_excel_file)